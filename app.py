import os
from io import BytesIO
from typing import Dict, List, Optional

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook


# ============================================================
# STEP 2 FINAL - ZSD50 FUEL + ZSD50G LPG/CNG P&L MAPPING
#
# Bu kod iki dosyayı birlikte çalıştırır:
# 1) ZSD50 Fuel
# 2) ZSD50G LPG/CNG
#
# Admin dosyaları yükler, uygulama bunları tek master P&L formatına çevirir.
# Normal kullanıcı sadece dashboard görür.
#
# ZSD50 Fuel mapping:
# - Header row: 2
# - Ftrl.mkt.        -> Fuel Volume
# - Satış Fiyatı     -> Fuel Gross Sales
# - İndirim          -> Discount Included in Invoice Total
# - Katkı,Opr.       -> Additive
# - Toplam Maliyet   -> Fuel COGS
#
# ZSD50G LPG/CNG mapping:
# - Header row: 1
# - Ağırlık(*)       -> Gas Volume
# - Satış Fiya       -> Gas Gross Sales
# - İndirim          -> Discount Included in Invoice Total
# - TL MAliyet       -> Gas COGS
# - Net değer        -> net satış kontrolü
#
# P&L'de LPG/CNG satırları "Gas Volume / Gas Gross Sales / Gas COGS"
# olarak tutulur. Dashboardda LPG/CNG olarak gösterilir.
# ============================================================

st.set_page_config(
    page_title="Fuel + LPG/CNG P&L Mapping",
    page_icon="⛽",
    layout="wide",
)

DATA_DIR = "pnl_app_data"
MASTER_FILE = os.path.join(DATA_DIR, "fuel_lpg_cng_master.csv")
os.makedirs(DATA_DIR, exist_ok=True)

try:
    ADMIN_PASSWORD = st.secrets.get("ADMIN_PASSWORD", "admin123")
except Exception:
    ADMIN_PASSWORD = "admin123"


# ------------------------------------------------------------
# Basic helpers
# ------------------------------------------------------------

def clean_col(col) -> str:
    return str(col).replace("\u00a0", " ").strip()


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [clean_col(c) for c in df.columns]

    seen = {}
    new_cols = []

    for c in df.columns:
        if c == "" or c.lower().startswith("none"):
            c = "Unnamed"

        if c not in seen:
            seen[c] = 0
            new_cols.append(c)
        else:
            seen[c] += 1
            new_cols.append(f"{c}_{seen[c]}")

    df.columns = new_cols

    # Remove fully empty columns
    df = df.dropna(axis=1, how="all")
    return df


def normalize_text_for_match(x: str) -> str:
    return (
        clean_col(x)
        .lower()
        .replace("ı", "i")
        .replace("İ", "i")
        .replace("ğ", "g")
        .replace("Ğ", "g")
        .replace("ü", "u")
        .replace("Ü", "u")
        .replace("ş", "s")
        .replace("Ş", "s")
        .replace("ö", "o")
        .replace("Ö", "o")
        .replace("ç", "c")
        .replace("Ç", "c")
    )


def find_col(df: pd.DataFrame, aliases: List[str], required: bool = True) -> Optional[str]:
    cols = list(df.columns)
    exact_map = {clean_col(c).lower(): c for c in cols}
    norm_map = {normalize_text_for_match(c): c for c in cols}

    # Exact lowercase match
    for alias in aliases:
        key = clean_col(alias).lower()
        if key in exact_map:
            return exact_map[key]

    # Turkish-insensitive exact match
    for alias in aliases:
        key = normalize_text_for_match(alias)
        if key in norm_map:
            return norm_map[key]

    # Contains match
    for alias in aliases:
        key = normalize_text_for_match(alias)
        for c in cols:
            if key in normalize_text_for_match(c):
                return c

    if required:
        raise ValueError(f"Zorunlu kolon bulunamadı. Aranan kolonlar: {aliases}")
    return None


def to_num(s: pd.Series) -> pd.Series:
    if pd.api.types.is_numeric_dtype(s):
        return pd.to_numeric(s, errors="coerce")

    raw = s.astype(str).str.strip()

    # Handles Turkish number style: 1.234.567,89
    cleaned = (
        raw
        .str.replace("\u00a0", "", regex=False)
        .str.replace(" ", "", regex=False)
        .str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False)
    )

    return pd.to_numeric(cleaned, errors="coerce")


def to_str(s: pd.Series) -> pd.Series:
    return (
        s.astype(str)
        .replace({"nan": "", "NaT": "", "<NA>": "", "None": ""})
        .str.strip()
    )


def get_sheet_names(uploaded_file) -> List[str]:
    lower = uploaded_file.name.lower()

    if lower.endswith((".xlsx", ".xlsm")):
        wb = load_workbook(BytesIO(uploaded_file.getvalue()), read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names

    if lower.endswith(".xls"):
        return pd.ExcelFile(BytesIO(uploaded_file.getvalue())).sheet_names

    return ["CSV"]


def read_excel_fast(
    uploaded_file,
    sheet_name: Optional[str],
    header_excel_row: int,
    max_excel_cols: int = 80,
) -> pd.DataFrame:
    """
    ZSD50G dosyasında Excel'in max_column değeri çok büyük görünebiliyor.
    Bu yüzden pandas read_excel yerine openpyxl ile sadece ilk 80 kolon okunur.
    """
    wb = load_workbook(BytesIO(uploaded_file.getvalue()), read_only=True, data_only=True)
    ws = wb[sheet_name or wb.sheetnames[0]]

    header_row = max(int(header_excel_row), 1)
    rows = ws.iter_rows(min_row=header_row, max_col=max_excel_cols, values_only=True)

    try:
        header_values = next(rows)
    except StopIteration:
        wb.close()
        raise ValueError("Excel sayfası boş görünüyor.")

    headers = []
    for i, value in enumerate(header_values, start=1):
        if value is None or clean_col(value) == "":
            headers.append(f"Unnamed_{i}")
        else:
            headers.append(clean_col(value))

    data = []
    for row in rows:
        if row is None:
            continue

        if all(v is None for v in row):
            continue

        data.append(list(row))

    wb.close()

    df = pd.DataFrame(data, columns=headers)
    return normalize_columns(df)


def read_file(
    uploaded_file,
    sheet_name: Optional[str],
    header_excel_row: int,
    max_excel_cols: int = 80,
) -> pd.DataFrame:
    lower = uploaded_file.name.lower()

    if lower.endswith(".csv"):
        header_idx = max(int(header_excel_row) - 1, 0)
        df = pd.read_csv(BytesIO(uploaded_file.getvalue()), header=header_idx)
        return normalize_columns(df)

    if lower.endswith((".xlsx", ".xlsm")):
        return read_excel_fast(
            uploaded_file=uploaded_file,
            sheet_name=sheet_name,
            header_excel_row=header_excel_row,
            max_excel_cols=max_excel_cols,
        )

    if lower.endswith(".xls"):
        header_idx = max(int(header_excel_row) - 1, 0)
        df = pd.read_excel(BytesIO(uploaded_file.getvalue()), sheet_name=sheet_name, header=header_idx)
        return normalize_columns(df)

    raise ValueError("Sadece CSV / XLSX / XLSM / XLS dosyası desteklenir.")


# ------------------------------------------------------------
# Source-specific mappings
# ------------------------------------------------------------

SOURCE_FUEL = "ZSD50 Fuel"
SOURCE_GAS = "ZSD50G LPG/CNG"


def source_to_prefix(source_type: str) -> str:
    return "Gas" if source_type == SOURCE_GAS else "Fuel"


def detect_sales_columns(df: pd.DataFrame, source_type: str) -> Dict[str, Optional[str]]:
    """
    ZSD50 ve ZSD50G kolonları aynı değil.
    Bu yüzden bazı alanlarda kaynak tipine göre öncelikli alias listesi kullanıyoruz.
    """

    if source_type == SOURCE_GAS:
        return {
            "date": find_col(df, ["Ftrl.trh.", "Fatura Tarihi", "Date", "Tarih"]),
            "station": find_col(df, ["İstasyon", "Istasyon", "Station", "Name&TrCode"]),
            "name": find_col(df, ["Name&TrCode", "Name Retrieve", "Ad 1", "Name"], required=False),
            "dealer_type": find_col(df, ["Firma", "Satış tipi", "Satis tipi", "Kanal", "Counterparty"], required=False),
            "supply_city": find_col(df, ["Br.Tanım", "Br.Tanim", "Supply city for fuel", "Supply City"], required=False),
            "material": find_col(df, ["Malzeme", "Material"], required=False),
            "description": find_col(df, ["Tanım", "Tanim", "Ürün Açıklaması", "Urun Aciklamasi"], required=False),
            "product": find_col(df, ["Ürün Grubu", "Urun Grubu", "Product", "Tanım", "Tanim"], required=False),
            "volume": find_col(df, ["Ağırlık(*)", "Agirlik(*)", "Ağırlık", "Agirlik", "Ftrl.mkt.", "Volume", "Miktar"]),
            "gross_sales": find_col(df, ["Satış Fiya", "Satis Fiya", "Satış Fiyatı", "Satis Fiyati", "Gross Sales", "Net değer"]),
            "discount": find_col(df, ["İndirim", "Indirim", "Discount"], required=False),
            "net_value": find_col(df, ["Net değer", "Net deger", "TL-Net Değer", "TL-Net Deger"], required=False),
            "additive": find_col(df, ["Katkı,Opr.", "Katki,Opr.", "Additive"], required=False),
            "cogs": find_col(df, ["TL MAliyet", "TL Maliyet", "TL-Maliyet", "Toplam Maliyet", "Total Cost", "COGS", "Maliyet"]),
            "rate": find_col(df, ["Rate", "Kur"], required=False),
            "usd_net_value": find_col(df, ["$  Net değer", "$ Net değer", "$ Net deger", "USD NET DEĞER", "USD NET DEGER"], required=False),
            "usd_cogs": find_col(df, ["$ Maliyet", "USD MALİYET", "USD MALIYET", "USD Cost"], required=False),
            "invoice": find_col(df, ["Ftr.Matbu No", "Ftr.blg", "Fatura No"], required=False),
        }

    return {
        "date": find_col(df, ["Ftrl.trh.", "Fatura Tarihi", "Date", "Tarih"]),
        "station": find_col(df, ["İstasyon", "Istasyon", "Station", "TR Code", "Name&TrCode"]),
        "name": find_col(df, ["Name&TrCode", "Name Retrieve", "Ad 1", "Name", "Bayi Adı"], required=False),
        "dealer_type": find_col(df, ["Kanal", "Counterparty", "Dealer/Acenta", "Dealer_Acenta", "Satış tipi"], required=False),
        "supply_city": find_col(df, ["Br.Tanım", "Br.Tanim", "Supply city for fuel", "Supply City"], required=False),
        "material": find_col(df, ["Malzeme", "Material"], required=False),
        "description": find_col(df, ["Tanım", "Tanim", "Ürün Açıklaması", "Urun Aciklamasi"], required=False),
        "product": find_col(df, ["Product", "Ürün Açıklaması", "Urun Aciklamasi", "Tanım", "Tanim"], required=False),
        "volume": find_col(df, ["Ftrl.mkt.", "Ftrl.mkt", "Faturalanan miktar", "Volume", "Miktar"]),
        "gross_sales": find_col(df, ["Satış Fiyatı", "Satis Fiyati", "Gross Sales", "Brüt Satış", "Brut Satis"]),
        "discount": find_col(df, ["İndirim", "Indirim", "Discount"], required=False),
        "net_value": find_col(df, ["TL-Net Değer", "TL-Net Deger", "Net Değer", "Net Deger"], required=False),
        "additive": find_col(df, ["Katkı,Opr.", "Katki,Opr.", "Additive"], required=False),
        "cogs": find_col(df, ["Toplam Maliyet", "Total Cost", "COGS", "Maliyet"]),
        "rate": find_col(df, ["Rate", "Kur"], required=False),
        "usd_net_value": find_col(df, ["USD NET DEĞER", "USD NET DEGER", "USD Net"], required=False),
        "usd_cogs": find_col(df, ["USD MALİYET", "USD MALIYET", "USD Cost"], required=False),
        "invoice": find_col(df, ["Ftr.Matbu No", "Ftr.blg", "Fatura No"], required=False),
    }


def classify_product(product: str, description: str, source_type: str) -> str:
    text = f"{product} {description}".upper()

    if source_type == SOURCE_GAS:
        if "CNG" in text:
            return "CNG"
        if "LNG" in text:
            return "LNG"
        if "LPG" in text or "OTOGAZ" in text or "AUTOGAS" in text or "AUTO GAS" in text:
            return "LPG"
        return "LPG/CNG"

    if "LPG" in text or "OTOGAZ" in text or "AUTOGAS" in text:
        return "LPG"
    if "GASOLINE" in text or "BENZ" in text or "KURŞUNSUZ" in text or "KURSUNSUZ" in text:
        return "Gasoline"
    if "DIESEL" in text or "MOTORIN" in text or "MOTORİN" in text or "ECTO" in text:
        return "Diesel"

    return "Other"


def amount_tl(df: pd.DataFrame, col: Optional[str]) -> pd.Series:
    if col is None or col not in df.columns:
        return pd.Series(0.0, index=df.index)
    return to_num(df[col]).fillna(0)


def amount_usd_from_tl(df: pd.DataFrame, tl_col: Optional[str], rate_col: Optional[str]) -> pd.Series:
    if tl_col is None or tl_col not in df.columns:
        return pd.Series(0.0, index=df.index)

    if rate_col is None or rate_col not in df.columns:
        return pd.Series(0.0, index=df.index)

    tl = to_num(df[tl_col]).fillna(0)
    rate = to_num(df[rate_col]).replace(0, np.nan)
    return (tl / rate).fillna(0)


def build_line(base: pd.DataFrame, line: str, value: pd.Series, currency: str) -> pd.DataFrame:
    out = base.copy()
    out["P&L Line"] = line
    out["Currency"] = currency
    out["Value"] = pd.to_numeric(value, errors="coerce").fillna(0)
    return out


def map_sales_file_to_pnl(
    df: pd.DataFrame,
    source_type: str,
    make_cogs_negative: bool = True,
    include_usd: bool = True,
) -> pd.DataFrame:
    cols = detect_sales_columns(df, source_type)

    valid = df.copy()
    valid[cols["date"]] = pd.to_datetime(valid[cols["date"]], errors="coerce")

    valid = valid[
        valid[cols["date"]].notna()
        & valid[cols["station"]].notna()
        & valid[cols["volume"]].notna()
    ].copy()

    if valid.empty:
        raise ValueError(
            "Geçerli satış satırı bulunamadı. Header satırını kontrol et. "
            "ZSD50 için genelde header 2, ZSD50G için header 1 olmalı."
        )

    prefix = source_to_prefix(source_type)

    base = pd.DataFrame(index=valid.index)
    base["Date"] = valid[cols["date"]]
    base["Year"] = base["Date"].dt.year
    base["Quarter"] = "Q" + base["Date"].dt.quarter.astype(str)
    base["Month"] = base["Date"].dt.strftime("%Y-%m")
    base["Day"] = base["Date"].dt.strftime("%Y-%m-%d")
    base["Year_Month_Quarter"] = (
        base["Year"].astype(str)
        + "-"
        + base["Quarter"].astype(str)
        + "-"
        + base["Date"].dt.month.astype(str)
    )

    base["İstasyon"] = to_str(valid[cols["station"]])
    base["Name"] = to_str(valid[cols["name"]]) if cols["name"] else ""
    base["Dealer/Acenta"] = to_str(valid[cols["dealer_type"]]) if cols["dealer_type"] else ""
    base["Supply city for fuel"] = to_str(valid[cols["supply_city"]]) if cols["supply_city"] else ""

    product_series = to_str(valid[cols["product"]]) if cols["product"] else ""
    desc_series = to_str(valid[cols["description"]]) if cols["description"] else ""

    if isinstance(product_series, str):
        product_series = pd.Series(product_series, index=valid.index)
    if isinstance(desc_series, str):
        desc_series = pd.Series(desc_series, index=valid.index)

    base["Product"] = product_series
    base["Product Description"] = desc_series
    base["Product Group"] = [
        classify_product(p, d, source_type)
        for p, d in zip(product_series.tolist(), desc_series.tolist())
    ]

    base["Source"] = source_type
    base["Raw Rows"] = 1

    if cols["invoice"]:
        base["Invoice"] = to_str(valid[cols["invoice"]])
    else:
        base["Invoice"] = ""

    # TL values
    volume = amount_tl(valid, cols["volume"])
    gross_sales = amount_tl(valid, cols["gross_sales"])
    discount = amount_tl(valid, cols["discount"])
    additive = amount_tl(valid, cols["additive"])

    cogs_raw = amount_tl(valid, cols["cogs"])
    cogs = -cogs_raw.abs() if make_cogs_negative else cogs_raw

    net_file = amount_tl(valid, cols["net_value"])
    calculated_net = gross_sales + discount + additive
    diff = calculated_net - net_file

    result = [
        build_line(base, f"{prefix} Volume", volume, "TL"),
        build_line(base, f"{prefix} Gross Sales", gross_sales, "TL"),
        build_line(base, "Discount Included in Invoice Total", discount, "TL"),
        build_line(base, "Additive", additive, "TL"),
        build_line(base, f"{prefix} COGS", cogs, "TL"),

        # Total rows. When Fuel + LPG/CNG are both loaded, these aggregate naturally.
        build_line(base, "Volume Total", volume, "TL"),
        build_line(base, "Gross Sales Total", gross_sales, "TL"),
        build_line(base, "COGS Total", cogs, "TL"),
        build_line(base, "Gross Margin", gross_sales + discount + additive + cogs, "TL"),

        # Control rows
        build_line(base, "CONTROL - Net Sales From File", net_file, "TL"),
        build_line(base, "CONTROL - Calculated Net Sales", calculated_net, "TL"),
        build_line(base, "CONTROL - Net Sales Difference", diff, "TL"),
    ]

    if include_usd:
        gross_sales_usd = amount_usd_from_tl(valid, cols["gross_sales"], cols["rate"])
        discount_usd = amount_usd_from_tl(valid, cols["discount"], cols["rate"])
        additive_usd = amount_usd_from_tl(valid, cols["additive"], cols["rate"])

        if cols["usd_cogs"] and cols["usd_cogs"] in valid.columns:
            cogs_usd_raw = amount_tl(valid, cols["usd_cogs"])
        else:
            cogs_usd_raw = amount_usd_from_tl(valid, cols["cogs"], cols["rate"])

        cogs_usd = -cogs_usd_raw.abs() if make_cogs_negative else cogs_usd_raw

        if cols["usd_net_value"] and cols["usd_net_value"] in valid.columns:
            net_file_usd = amount_tl(valid, cols["usd_net_value"])
        else:
            net_file_usd = amount_usd_from_tl(valid, cols["net_value"], cols["rate"])

        calculated_net_usd = gross_sales_usd + discount_usd + additive_usd
        diff_usd = calculated_net_usd - net_file_usd

        result.extend([
            build_line(base, f"{prefix} Volume", volume, "USD"),
            build_line(base, f"{prefix} Gross Sales", gross_sales_usd, "USD"),
            build_line(base, "Discount Included in Invoice Total", discount_usd, "USD"),
            build_line(base, "Additive", additive_usd, "USD"),
            build_line(base, f"{prefix} COGS", cogs_usd, "USD"),

            build_line(base, "Volume Total", volume, "USD"),
            build_line(base, "Gross Sales Total", gross_sales_usd, "USD"),
            build_line(base, "COGS Total", cogs_usd, "USD"),
            build_line(base, "Gross Margin", gross_sales_usd + discount_usd + additive_usd + cogs_usd, "USD"),

            build_line(base, "CONTROL - Net Sales From File", net_file_usd, "USD"),
            build_line(base, "CONTROL - Calculated Net Sales", calculated_net_usd, "USD"),
            build_line(base, "CONTROL - Net Sales Difference", diff_usd, "USD"),
        ])

    return pd.concat(result, ignore_index=True)


# ------------------------------------------------------------
# Master save/load
# ------------------------------------------------------------

def save_master(df: pd.DataFrame) -> None:
    df.to_csv(MASTER_FILE, index=False, encoding="utf-8-sig")
    st.session_state["master_df"] = df.copy()


def load_master() -> pd.DataFrame:
    if "master_df" in st.session_state and not st.session_state["master_df"].empty:
        return st.session_state["master_df"].copy()

    if os.path.exists(MASTER_FILE):
        df = pd.read_csv(MASTER_FILE)
        if "Date" in df.columns:
            df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        return df

    return pd.DataFrame()


def clear_master() -> None:
    if os.path.exists(MASTER_FILE):
        os.remove(MASTER_FILE)

    if "master_df" in st.session_state:
        del st.session_state["master_df"]


def upsert_source(new_source_df: pd.DataFrame, source_type: str, replace_same_source: bool = True) -> pd.DataFrame:
    master = load_master()

    if master.empty:
        combined = new_source_df.copy()
    else:
        if replace_same_source and "Source" in master.columns:
            master = master[master["Source"].astype(str) != source_type].copy()
        combined = pd.concat([master, new_source_df], ignore_index=True)

    save_master(combined)
    return combined


# ------------------------------------------------------------
# Analytics helpers
# ------------------------------------------------------------

def aggregate(df: pd.DataFrame, group_cols: List[str]) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()

    group_cols = [c for c in group_cols if c in df.columns]
    return df.groupby(group_cols + ["P&L Line"], dropna=False, as_index=False)["Value"].sum()


def pivot_pnl(df: pd.DataFrame, index_cols: List[str]) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()

    index_cols = [c for c in index_cols if c in df.columns]

    piv = df.pivot_table(
        index=index_cols,
        columns="P&L Line",
        values="Value",
        aggfunc="sum",
        fill_value=0,
    ).reset_index()

    piv.columns.name = None

    needed = [
        "Fuel Volume",
        "Gas Volume",
        "Volume Total",
        "Fuel Gross Sales",
        "Gas Gross Sales",
        "Gross Sales Total",
        "Discount Included in Invoice Total",
        "Additive",
        "Fuel COGS",
        "Gas COGS",
        "COGS Total",
        "Gross Margin",
        "CONTROL - Net Sales Difference",
    ]

    for col in needed:
        if col not in piv.columns:
            piv[col] = 0.0

    piv["Gross Margin %"] = np.where(
        piv["Gross Sales Total"] != 0,
        piv["Gross Margin"] / piv["Gross Sales Total"],
        np.nan,
    )

    piv["Gross Margin / Volume"] = np.where(
        piv["Volume Total"] != 0,
        piv["Gross Margin"] / piv["Volume Total"],
        np.nan,
    )

    return piv


def money(v, currency):
    symbol = "₺" if currency == "TL" else "$"
    if pd.isna(v):
        v = 0
    return f"{symbol}{v:,.0f}"


def number(v):
    if pd.isna(v):
        v = 0
    return f"{v:,.0f}"


def pct(v):
    if pd.isna(v):
        return "-"
    return f"{v:.2%}"


def show_dashboard(master: pd.DataFrame) -> None:
    if master.empty:
        st.info("Henüz master rapor yok. Admin Panel’den ZSD50 Fuel ve/veya ZSD50G LPG/CNG dosyasını yükle.")
        return

    st.subheader("📊 Fuel + LPG/CNG P&L Dashboard")

    st.sidebar.header("Dashboard filtreleri")

    currencies = sorted(master["Currency"].dropna().unique().tolist())
    default_currency_index = currencies.index("TL") if "TL" in currencies else 0
    currency = st.sidebar.selectbox("Para birimi", currencies, index=default_currency_index)

    df = master[master["Currency"].eq(currency)].copy()

    if "Source" in df.columns:
        sources = sorted(df["Source"].dropna().unique().tolist())
        selected_sources = st.sidebar.multiselect("Kaynak", sources, default=sources)
        if selected_sources:
            df = df[df["Source"].isin(selected_sources)]

    months = sorted(df["Month"].dropna().unique().tolist())
    selected_months = st.sidebar.multiselect("Ay", months, default=months)
    if selected_months:
        df = df[df["Month"].isin(selected_months)]

    dealer_types = sorted(df["Dealer/Acenta"].dropna().unique().tolist())
    selected_dealer_types = st.sidebar.multiselect("Dealer/Acenta", dealer_types, default=dealer_types)
    if selected_dealer_types:
        df = df[df["Dealer/Acenta"].isin(selected_dealer_types)]

    product_groups = sorted(df["Product Group"].dropna().unique().tolist())
    selected_product_groups = st.sidebar.multiselect("Ürün grubu", product_groups, default=product_groups)
    if selected_product_groups:
        df = df[df["Product Group"].isin(selected_product_groups)]

    station_search = st.sidebar.text_input("İstasyon / bayi ara", "")
    if station_search:
        mask = (
            df["İstasyon"].astype(str).str.contains(station_search, case=False, na=False)
            | df["Name"].astype(str).str.contains(station_search, case=False, na=False)
        )
        df = df[mask]

    financial = df[~df["P&L Line"].str.startswith("CONTROL", na=False)].copy()
    control = df[df["P&L Line"].str.startswith("CONTROL", na=False)].copy()

    total_long = aggregate(financial, ["Currency"])
    total = pivot_pnl(total_long, ["Currency"])

    if total.empty:
        st.warning("Seçili filtrelerde finansal veri yok.")
        return

    r = total.iloc[0]

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Total Volume", number(r.get("Volume Total", 0)))
    c2.metric("Fuel Volume", number(r.get("Fuel Volume", 0)))
    c3.metric("LPG/CNG Volume", number(r.get("Gas Volume", 0)))
    c4.metric("Gross Sales", money(r.get("Gross Sales Total", 0), currency))
    c5.metric("COGS", money(r.get("COGS Total", 0), currency))
    c6.metric("Gross Margin", money(r.get("Gross Margin", 0), currency))

    m1, m2, m3 = st.columns(3)
    m1.metric("Gross Margin %", pct(r.get("Gross Margin %", np.nan)))
    m2.metric(
        "GM / Volume",
        "-" if pd.isna(r.get("Gross Margin / Volume", np.nan)) else f"{r.get('Gross Margin / Volume'):,.4f}",
    )
    m3.metric("Discount", money(r.get("Discount Included in Invoice Total", 0), currency))

    tab1, tab2, tab3, tab4, tab5 = st.tabs(
        ["İstasyon Özeti", "Kaynak / Ürün Özeti", "Ay Trend", "P&L Statement", "Kontrol"]
    )

    with tab1:
        st.markdown("#### İstasyon bazında Fuel + LPG/CNG")
        station_long = aggregate(financial, ["İstasyon", "Name", "Dealer/Acenta"])
        station_summary = pivot_pnl(station_long, ["İstasyon", "Name", "Dealer/Acenta"])
        station_summary = station_summary.sort_values("Gross Margin", ascending=False)

        st.dataframe(station_summary, use_container_width=True, height=620)

        csv = station_summary.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
        st.download_button("İstasyon özetini CSV indir", csv, "fuel_lpg_cng_station_summary.csv", "text/csv")

    with tab2:
        col1, col2 = st.columns(2)

        with col1:
            st.markdown("#### Kaynak bazında")
            source_long = aggregate(financial, ["Source"])
            source_summary = pivot_pnl(source_long, ["Source"]).sort_values("Gross Margin", ascending=False)
            st.dataframe(source_summary, use_container_width=True)

        with col2:
            st.markdown("#### Ürün grubu bazında")
            product_long = aggregate(financial, ["Product Group"])
            product_summary = pivot_pnl(product_long, ["Product Group"]).sort_values("Gross Margin", ascending=False)
            st.dataframe(product_summary, use_container_width=True)

    with tab3:
        st.markdown("#### Ay bazında trend")
        month_long = aggregate(financial, ["Month"])
        month_summary = pivot_pnl(month_long, ["Month"]).sort_values("Month")
        st.dataframe(month_summary, use_container_width=True)

        if not month_summary.empty:
            st.line_chart(month_summary.set_index("Month")[["Gross Sales Total", "COGS Total", "Gross Margin"]])

    with tab4:
        st.markdown("#### P&L statement")
        order = [
            "Fuel Volume",
            "Gas Volume",
            "Volume Total",
            "Fuel Gross Sales",
            "Gas Gross Sales",
            "Gross Sales Total",
            "Discount Included in Invoice Total",
            "Additive",
            "Fuel COGS",
            "Gas COGS",
            "COGS Total",
            "Gross Margin",
        ]
        order_map = {x: i for i, x in enumerate(order)}

        statement = financial.groupby("P&L Line", as_index=False)["Value"].sum()
        statement["Order"] = statement["P&L Line"].map(order_map).fillna(999).astype(int)
        statement = statement.sort_values(["Order", "P&L Line"])

        st.dataframe(statement[["P&L Line", "Value"]], use_container_width=True)

    with tab5:
        st.markdown("#### Net satış kontrolü")
        st.write(
            "Kontrol: `Gross Sales + Discount + Additive` ile dosyadaki net değer karşılaştırılır. "
            "Fark sıfıra yakın olmalı."
        )

        control_long = aggregate(control, ["Currency", "Source"])
        control_summary = pivot_pnl(control_long, ["Currency", "Source"])
        st.dataframe(control_summary, use_container_width=True)

        if not control_summary.empty:
            diff = control_summary["CONTROL - Net Sales Difference"].sum()
            st.metric("Net satış kontrol farkı", money(diff, currency))

        st.markdown("#### Master long format örneği")
        st.dataframe(df.head(500), use_container_width=True, height=400)


# ------------------------------------------------------------
# Admin upload UI
# ------------------------------------------------------------

def admin_upload_block(
    title: str,
    source_type: str,
    key_prefix: str,
    default_header_row: int,
) -> None:
    st.markdown(f"### {title}")

    uploaded = st.file_uploader(
        f"{title} dosyasını yükle",
        type=["xlsx", "xlsm", "xls", "csv"],
        key=f"{key_prefix}_uploader",
    )

    if uploaded is None:
        st.info(f"{title} için dosya yüklenmedi.")
        return

    sheet_name = None
    if uploaded.name.lower().endswith((".xlsx", ".xlsm", ".xls")):
        sheets = get_sheet_names(uploaded)
        default_sheet_index = 0

        # Uploaded samples use ZSD50 / ZSD50G sheet names
        for i, s in enumerate(sheets):
            if source_type == SOURCE_FUEL and "ZSD50" in str(s).upper() and "G" not in str(s).upper():
                default_sheet_index = i
                break
            if source_type == SOURCE_GAS and "ZSD50G" in str(s).upper():
                default_sheet_index = i
                break

        sheet_name = st.selectbox("Excel sayfası", sheets, index=default_sheet_index, key=f"{key_prefix}_sheet")

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        header_row = st.number_input(
            "Header Excel satırı",
            min_value=1,
            max_value=20,
            value=default_header_row,
            step=1,
            key=f"{key_prefix}_header",
        )
    with c2:
        max_excel_cols = st.number_input(
            "Okunacak maksimum Excel kolonu",
            min_value=20,
            max_value=200,
            value=80,
            step=10,
            key=f"{key_prefix}_maxcols",
            help="ZSD50G dosyasında boş formatlı kolonlar çok fazla görünebilir; bu limit dosyayı hızlı okutur.",
        )
    with c3:
        make_cogs_negative = st.checkbox("Maliyeti negatif yaz", value=True, key=f"{key_prefix}_cogs")
    with c4:
        include_usd = st.checkbox("USD üret", value=True, key=f"{key_prefix}_usd")

    replace_same_source = st.checkbox(
        "Aynı kaynağı master içinde değiştir; duplicate oluşturma",
        value=True,
        key=f"{key_prefix}_replace",
    )

    try:
        raw = read_file(
            uploaded_file=uploaded,
            sheet_name=sheet_name,
            header_excel_row=header_row,
            max_excel_cols=max_excel_cols,
        )

        detected = detect_sales_columns(raw, source_type)

        st.markdown("#### 1) Dosya okuma kontrolü")
        st.write(f"Okunan satır: **{len(raw):,}** | Okunan kolon: **{len(raw.columns):,}**")
        st.dataframe(raw.head(8), use_container_width=True)

        st.markdown("#### 2) Otomatik mapping kontrolü")
        mapping_df = pd.DataFrame(
            [{"Alan": k, "Bulunan kolon": v} for k, v in detected.items()]
        )
        st.dataframe(mapping_df, use_container_width=True)

        if st.button(f"✅ {title} master'a ekle / güncelle", type="primary", key=f"{key_prefix}_button"):
            mapped = map_sales_file_to_pnl(
                raw,
                source_type=source_type,
                make_cogs_negative=make_cogs_negative,
                include_usd=include_usd,
            )

            combined = upsert_source(
                new_source_df=mapped,
                source_type=source_type,
                replace_same_source=replace_same_source,
            )

            st.success(
                f"{title} master'a eklendi/güncellendi. "
                f"Bu kaynak P&L satırı: {len(mapped):,} | Toplam master P&L satırı: {len(combined):,}"
            )

            st.markdown("#### Bu kaynaktan oluşan P&L örneği")
            st.dataframe(mapped.head(100), use_container_width=True)

            st.markdown("### Güncel dashboard önizlemesi")
            show_dashboard(combined)

    except Exception as exc:
        st.error(f"Hata: {exc}")


# ------------------------------------------------------------
# Main UI
# ------------------------------------------------------------

st.title("⛽ Step 2 Final - ZSD50 Fuel + ZSD50G LPG/CNG")
st.caption("Fuel ve LPG/CNG satış raporlarını tek master P&L formatında birleştirir.")

with st.expander("Mapping özeti", expanded=True):
    st.markdown(
        """
        **ZSD50 Fuel**
        - Header satırı: `2`
        - `Ftrl.mkt.` → `Fuel Volume`
        - `Satış Fiyatı` → `Fuel Gross Sales`
        - `İndirim` → `Discount Included in Invoice Total`
        - `Katkı,Opr.` → `Additive`
        - `Toplam Maliyet` → `Fuel COGS`

        **ZSD50G LPG/CNG**
        - Header satırı: `1`
        - `Ağırlık(*)` → `Gas Volume`
        - `Satış Fiya` → `Gas Gross Sales`
        - `İndirim` → `Discount Included in Invoice Total`
        - `TL MAliyet` → `Gas COGS`
        - `Net değer` → kontrol amaçlı net satış

        P&L tarafında LPG/CNG satırları şirket şemasına uyumlu olması için
        `Gas Volume`, `Gas Gross Sales`, `Gas COGS` olarak tutulur.
        """
    )

screen = st.sidebar.radio("Ekran", ["Dashboard", "Admin Panel"], index=0)

if screen == "Admin Panel":
    st.subheader("🔐 Admin Panel")

    password = st.text_input("Admin şifresi", type="password")
    if password != ADMIN_PASSWORD:
        st.warning("Admin şifresi girildikten sonra upload alanları açılır.")
        st.stop()

    master = load_master()

    top1, top2, top3 = st.columns(3)
    with top1:
        if st.button("🗑️ Tüm master raporu temizle"):
            clear_master()
            st.success("Master temizlendi.")
            st.stop()

    with top2:
        if not master.empty:
            st.success(f"Mevcut master var: {len(master):,} P&L satırı")
        else:
            st.info("Mevcut master yok.")

    with top3:
        if not master.empty and "Source" in master.columns:
            sources = ", ".join(sorted(master["Source"].dropna().unique().tolist()))
            st.write(f"Kaynaklar: {sources}")

    if not master.empty:
        csv = master.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
        st.download_button("Master P&L CSV indir", csv, "fuel_lpg_cng_master.csv", "text/csv")

    st.divider()

    tab_fuel, tab_gas = st.tabs(["1) ZSD50 Fuel", "2) ZSD50G LPG/CNG"])

    with tab_fuel:
        admin_upload_block(
            title="ZSD50 Fuel",
            source_type=SOURCE_FUEL,
            key_prefix="fuel",
            default_header_row=2,
        )

    with tab_gas:
        admin_upload_block(
            title="ZSD50G LPG/CNG",
            source_type=SOURCE_GAS,
            key_prefix="gas",
            default_header_row=1,
        )

else:
    master = load_master()
    show_dashboard(master)
